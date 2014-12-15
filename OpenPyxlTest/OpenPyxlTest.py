from openpyxl import Workbook

# Create a workbook.
wb = Workbook()

# Create a first worksheet. This method always returns the first worksheet in the workbook.
ws0 = wb.active

print(ws0.title) #Output: "Sheet"

# Workbook.create_sheet() creates an empty worksheet at the end of the workbook.
ws1 = wb.create_sheet()

ws2 = wb.create_sheet()

print(ws1.title) #Output: "Sheet1"
print(ws2.title) #Output: "Sheet2"

# Workbook.create_sheet() creates an empty worksheet at the given position.
ws_pi = wb.create_sheet(3)
ws_pi.title = 'PI=3.141...'

print('\n After creating the sheet "{}" at position (3): '.format(ws_pi.title))
print(wb.get_sheet_names())

# Workbook.create_sheet() pushes other workbooks to the back if it has to squeeze 
# a new worksheet into a taken position
ws_primus = wb.create_sheet(1)
ws_primus.title = 'primus'

print('\n After creating the sheet "{}" at position (1): '.format(ws_primus.title))
print(wb.get_sheet_names())
